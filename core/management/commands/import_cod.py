"""Load causes of death by age band and sex from the DOSM workbook.

    python manage.py import_cod ~/Downloads/cod_2023.xlsx --year 2023

Why this exists: the earlier import read the workbook with pandas defaults,
which took row 0 (the Malay title banner) as the header. That produced the
one-row `causes_of_death` table holding nothing but the title. These sheets
need the header found explicitly, so this command does that.

Sheet layout, using "10.2 (2)" (ages 41-59, by sex) as the example:

    row 0-3   title banner, Malay then English
    row 5-6   column headers, Malay then English
    row 8     "Lelaki/ Male"      <- block marker in column A
    row 8-17    1. .. 10.  the ten principal causes
    row 18      "Keseluruhan sebab" (all causes) = group total
    row 20    "Perempuan/ Female" <- next block marker
    ...

    col A  block marker (only on the first row of each block)
    col B  rank, "1." .. "10."
    col C  cause name, English, medically certified
    col D  death count
    col E  percentage of all deaths in this group

Only the medically certified list is loaded. The non-medically certified
columns further right are published in Malay only.

Ages 0-14 (sheet 8.2) has no by-sex variant in the workbook, so it is skipped.
"""

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import CauseOfDeath, MortalityRecord

# Sheets named like "10.2 (2)" hold the by-sex breakdown.
BY_SEX_SHEET = re.compile(r'^(\d+)\.2 \(2\)$')

MALE_MARKER = 'lelaki'
FEMALE_MARKER = 'perempuan'
TOTAL_MARKER = 'keseluruhan sebab'

# Column offsets within a data row (0-indexed).
COL_MARKER, COL_RANK, COL_CAUSE, COL_COUNT, COL_PCT = 0, 1, 2, 3, 4


class Command(BaseCommand):
    help = 'Import DOSM causes of death by age band and sex from an .xlsx file'

    def add_arguments(self, parser):
        parser.add_argument('workbook', help='Path to cod_<year>.xlsx')
        parser.add_argument('--year', type=int, required=True,
                            help='Reference year of the publication')
        parser.add_argument('--location', default='Malaysia')
        parser.add_argument('--dry-run', action='store_true',
                            help='Parse and report without writing to the database')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError(
                'openpyxl is required. Install it with: pip install openpyxl'
            ) from exc

        path = Path(options['workbook']).expanduser()
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = []

        for sheet_name in workbook.sheetnames:
            if not BY_SEX_SHEET.match(sheet_name):
                continue

            sheet = workbook[sheet_name]
            age_band = self._age_band(sheet)
            if not age_band:
                self.stdout.write(
                    f'  skipped {sheet_name}: no age band in title')
                continue

            parsed = self._parse_sheet(sheet, age_band)
            rows.extend(parsed)
            self.stdout.write(
                f'  {sheet_name}: age band {age_band}, {len(parsed)} rows')

        workbook.close()

        if not rows:
            raise CommandError(
                'No by-sex cause sheets found. Is this the DOSM causes of '
                'death workbook?'
            )

        if options['dry_run']:
            self._report(rows)
            self.stdout.write(self.style.WARNING(
                f'\nDry run: parsed {len(rows)} rows, nothing written.'))
            return

        created = self._write(rows, options['year'], options['location'])
        self._report(rows)
        self.stdout.write(self.style.SUCCESS(
            f'\nImported {created} mortality rows for {options["year"]}.'))

    # -- parsing ---------------------------------------------------

    def _age_band(self, sheet):
        """Read the age band out of the English title line."""
        for row in sheet.iter_rows(min_row=1, max_row=6, max_col=4,
                                   values_only=True):
            for value in row:
                if not value:
                    continue
                text = str(value).lower()
                if 'aged' not in text:
                    continue
                if match := re.search(r'aged\s+(\d+)\s*-\s*(\d+)', text):
                    return f'{match.group(1)}-{match.group(2)}'
                if match := re.search(r'aged\s+(\d+)\s+years?\s+and\s+over', text):
                    return f'{match.group(1)}+'
        return None

    def _parse_sheet(self, sheet, age_band):
        """Walk the sheet, splitting on the Male / Female block markers."""
        rows = []
        sex = None
        block = []

        for raw in sheet.iter_rows(values_only=True):
            cells = list(raw) + [None] * (5 - len(raw))
            marker = str(cells[COL_MARKER] or '').strip().lower()
            cause = str(cells[COL_CAUSE] or '').strip()

            if MALE_MARKER in marker:
                sex, block = 'male', []
            elif FEMALE_MARKER in marker:
                sex, block = 'female', []

            if sex is None:
                continue

            # "Keseluruhan sebab" closes a block and carries its total.
            if TOTAL_MARKER in cause.lower():
                total = self._int(cells[COL_COUNT])
                for entry in block:
                    entry['group_total'] = total or 0
                    entry['age_band'] = age_band
                    entry['sex'] = sex
                    rows.append(entry)
                block, sex = [], None
                continue

            rank = self._rank(cells[COL_RANK])
            count = self._int(cells[COL_COUNT])
            pct = self._float(cells[COL_PCT])
            if rank is None or not cause or count is None or pct is None:
                continue

            block.append({
                'rank': rank,
                'cause': self._clean_cause(cause),
                'death_count': count,
                'share_pct': round(pct, 2),
            })

        return rows

    @staticmethod
    def _clean_cause(value):
        # Some cells carry embedded newlines from the bilingual layout.
        return re.sub(r'\s+', ' ', value).strip()

    @staticmethod
    def _rank(value):
        text = str(value or '').strip().rstrip('.')
        return int(text) if text.isdigit() else None

    @staticmethod
    def _int(value):
        try:
            return int(float(str(value).replace(',', '').strip()))
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _float(value):
        try:
            return float(str(value).replace(',', '').strip())
        except (TypeError, ValueError):
            return None

    # -- writing ---------------------------------------------------

    @transaction.atomic
    def _write(self, rows, year, location):
        # Replace this year's rows so re-running is safe.
        MortalityRecord.objects.filter(
            year=year, location=location, certification='medical'
        ).delete()

        causes = {}
        for name in {row['cause'] for row in rows}:
            causes[name], _ = CauseOfDeath.objects.get_or_create(cause_name=name)

        MortalityRecord.objects.bulk_create([
            MortalityRecord(
                cause=causes[row['cause']],
                year=year,
                location=location,
                age_band=row['age_band'],
                sex=row['sex'],
                certification='medical',
                rank=row['rank'],
                death_count=row['death_count'],
                share_pct=row['share_pct'],
                group_total=row['group_total'],
            )
            for row in rows
        ])
        return len(rows)

    def _report(self, rows):
        self.stdout.write('\nTop cause per group:')
        seen = set()
        for row in sorted(rows, key=lambda r: (r['age_band'], r['sex'], r['rank'])):
            key = (row['age_band'], row['sex'])
            if key in seen or row['rank'] != 1:
                continue
            seen.add(key)
            self.stdout.write(
                f'  {row["age_band"]:<6} {row["sex"]:<7} '
                f'{row["cause"][:44]:<46} {row["share_pct"]}%'
            )
